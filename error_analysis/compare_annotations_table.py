"""
Generate a per-sample annotation comparison table for the 38 validation samples.

Outputs:
  - annotation_comparison.xlsx  (color-coded Excel)
  - annotation_comparison.csv   (raw data)

Columns per error type: WT | Evelyn | GPT-4o
Color key:
  green  = 1 (annotated)
  white  = 0 (not annotated)
  orange = disagreement among three annotators

Usage:
    python compare_annotations_table.py
    python compare_annotations_table.py --pred test_gpt4o_validation_38.csv --gt error_analysis_common_agree.xlsx
"""

import argparse
import csv
import os
import sys

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ── Column definitions ───────────────────────────────────────────────────────

TYPES = [
    ("failed_parameter_extraction",           "Evelyn - failed parameter extraction",               "WT - failed parameter extraction"),
    ("verification_logic_error",              "Evelyn - verification logic errors ",                "WT - verification logic errors"),
    ("incorrect_formula_criteria",            "Evelyn - incorrect formula & criteria application",  "WT -incorrect formula & criteria application"),
    ("computation_error",                     "Evelyn - computation errors",                        "WT - computation errors"),
    ("omitted_calculation_process_or_result", "Evelyn - omitting the calculation process or result","WT - omitting the calculation process or result"),
    ("other_error",                           "Evelyn - other errors",                              "WT - other errors"),
]
SHORT_LABELS = ["PE", "VL", "FC", "CE", "OC", "OE"]
FULL_LABELS  = [
    "Failed Parameter\nExtraction",
    "Verification\nLogic Error",
    "Incorrect Formula\n& Criteria",
    "Computation\nError",
    "Omitted Calculation\nProcess/Result",
    "Other\nError",
]

# ── Styles ────────────────────────────────────────────────────────────────────

GREEN  = PatternFill("solid", fgColor="C6EFCE")
ORANGE = PatternFill("solid", fgColor="FFD966")
WHITE  = PatternFill("solid", fgColor="FFFFFF")
GREY   = PatternFill("solid", fgColor="F2F2F2")
HEADER = PatternFill("solid", fgColor="4472C4")
SUBHDR = PatternFill("solid", fgColor="9DC3E6")

THIN  = Side(style="thin",   color="BFBFBF")
MED   = Side(style="medium", color="595959")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BORDER  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

BOLD_WHITE = Font(bold=True, color="FFFFFF")
BOLD_DARK  = Font(bold=True, color="262626")
MONO       = Font(name="Courier New", size=10)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ── Data loading ─────────────────────────────────────────────────────────────

def load_gt(xlsx_path: str, valid_indices: set):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    gt_wt, gt_ev = {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        try:
            idx = int(float(str(d["index"])))
        except Exception:
            continue
        if valid_indices and idx not in valid_indices:
            continue
        gt_wt[idx] = [int(d.get(wc) or 0) for _, _, wc in TYPES]
        gt_ev[idx] = [int(d.get(ec) or 0) for _, ec, _ in TYPES]
    return gt_wt, gt_ev


def load_predictions(csv_path: str):
    preds, meta = {}, {}
    with open(csv_path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            try:
                idx = int(float(row["index"]))
            except Exception:
                continue
            preds[idx] = [int(row.get(t, 0) or 0) for t, _, _ in TYPES]
            meta[idx]  = {
                "true_label":      row.get("true_label", ""),
                "predicted_label": row.get("predicted_label", ""),
                "claim":           row.get("claim", ""),
            }
    return preds, meta


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_excel(idxs, gt_wt, gt_ev, preds, meta, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annotation Comparison"

    # ── Row 1: type group headers ─────────────────────────────────────────────
    # Fixed columns: A=index, B=true_label, C=predicted_label
    fixed = 3
    n_types = len(TYPES)

    ws.cell(1, 1, "Index").font      = BOLD_WHITE
    ws.cell(1, 1).fill               = HEADER
    ws.cell(1, 1).alignment          = CENTER
    ws.cell(1, 2, "True Label").font = BOLD_WHITE
    ws.cell(1, 2).fill               = HEADER
    ws.cell(1, 2).alignment          = CENTER
    ws.cell(1, 3, "Pred Label").font = BOLD_WHITE
    ws.cell(1, 3).fill               = HEADER
    ws.cell(1, 3).alignment          = CENTER

    for i, (label, full) in enumerate(zip(SHORT_LABELS, FULL_LABELS)):
        start_col = fixed + i * 3 + 1
        ws.merge_cells(start_row=1, start_column=start_col,
                       end_row=1,   end_column=start_col + 2)
        c = ws.cell(1, start_col, f"{label}: {full}")
        c.font      = BOLD_WHITE
        c.fill      = HEADER
        c.alignment = CENTER

    # ── Row 2: annotator sub-headers ──────────────────────────────────────────
    ws.cell(2, 1).value     = ""
    ws.cell(2, 2).value     = ""
    ws.cell(2, 3).value     = ""
    for col in range(1, 4):
        ws.cell(2, col).fill      = SUBHDR
        ws.cell(2, col).alignment = CENTER

    for i in range(n_types):
        for j, name in enumerate(["WT", "Evelyn", "GPT-4o"]):
            c = ws.cell(2, fixed + i * 3 + 1 + j, name)
            c.fill      = SUBHDR
            c.font      = BOLD_DARK
            c.alignment = CENTER

    # ── Data rows ─────────────────────────────────────────────────────────────
    for r, idx in enumerate(idxs, start=3):
        fill_row = GREY if r % 2 == 0 else WHITE

        ws.cell(r, 1, idx).alignment = CENTER
        ws.cell(r, 1).font           = MONO
        ws.cell(r, 2, meta[idx]["true_label"]).alignment      = CENTER
        ws.cell(r, 3, meta[idx]["predicted_label"]).alignment = CENTER
        for col in range(1, 4):
            ws.cell(r, col).fill = fill_row

        for i in range(n_types):
            wt_val = gt_wt.get(idx, [0]*n_types)[i]
            ev_val = gt_ev.get(idx, [0]*n_types)[i]
            gp_val = preds.get(idx, [0]*n_types)[i]
            vals   = [wt_val, ev_val, gp_val]
            disagree = len(set(vals)) > 1

            for j, v in enumerate(vals):
                col = fixed + i * 3 + 1 + j
                c = ws.cell(r, col, v)
                c.alignment = CENTER
                c.font      = MONO
                if disagree:
                    c.fill = ORANGE
                elif v == 1:
                    c.fill = GREEN
                else:
                    c.fill = fill_row

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    col_letters = "DEFGHIJKLMNOPQRSTUVWX"
    for i, ch in enumerate(col_letters[:n_types * 3]):
        ws.column_dimensions[ch].width = 8

    # ── Row heights ───────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 18

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = "D3"

    # ── Legend sheet ─────────────────────────────────────────────────────────
    lg = wb.create_sheet("Legend")
    legend_data = [
        ("Color",    "Meaning"),
        ("Green",    "Value = 1 (error annotated), all three agree"),
        ("White",    "Value = 0 (not annotated), all three agree"),
        ("Orange",   "Disagreement among WT / Evelyn / GPT-4o"),
    ]
    for i, (k, v) in enumerate(legend_data, 1):
        lg.cell(i, 1, k).font = Font(bold=(i == 1))
        lg.cell(i, 2, v)
    lg.cell(2, 1).fill = GREEN
    lg.cell(3, 1).fill = WHITE
    lg.cell(4, 1).fill = ORANGE
    lg.column_dimensions["A"].width = 10
    lg.column_dimensions["B"].width = 55

    # ── Error-type key sheet ──────────────────────────────────────────────────
    ek = wb.create_sheet("Error Type Key")
    ek.cell(1, 1, "Code").font = Font(bold=True)
    ek.cell(1, 2, "Full Name").font = Font(bold=True)
    for i, (s, full) in enumerate(zip(SHORT_LABELS, [t for t, _, _ in TYPES]), 2):
        ek.cell(i, 1, s)
        ek.cell(i, 2, full)
    ek.column_dimensions["A"].width = 6
    ek.column_dimensions["B"].width = 50

    wb.save(out_path)
    print(f"  Saved: {out_path}")


# ── CSV builder ────────────────────────────────────────────────────────────────

def build_csv(idxs, gt_wt, gt_ev, preds, meta, out_path):
    n_types = len(TYPES)
    header = ["index", "true_label", "pred_label"]
    for s, (t, _, _) in zip(SHORT_LABELS, TYPES):
        header += [f"{s}_WT", f"{s}_Evelyn", f"{s}_GPT4o"]

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for idx in idxs:
            row = [idx, meta[idx]["true_label"], meta[idx]["predicted_label"]]
            for i in range(n_types):
                row.append(gt_wt.get(idx, [0]*n_types)[i])
                row.append(gt_ev.get(idx, [0]*n_types)[i])
                row.append(preds.get(idx, [0]*n_types)[i])
            w.writerow(row)
    print(f"  Saved: {out_path}")


# ── Terminal table ────────────────────────────────────────────────────────────

def print_terminal(idxs, gt_wt, gt_ev, preds, meta):
    n_types = len(TYPES)
    # Header
    header1 = f"{'idx':>4}  {'True':>11}  {'Pred':>11}  "
    header1 += "  ".join(f"{'─'+s+'─':^9}" for s in SHORT_LABELS)
    header2 = f"{'':>4}  {'Label':>11}  {'Label':>11}  "
    header2 += "  ".join(f"{'W  E  G':^9}" for _ in SHORT_LABELS)
    sep = "─" * len(header1)

    print("\n" + sep)
    print(header1)
    print(header2)
    print(sep)

    for idx in idxs:
        tl = meta[idx]["true_label"][:11]
        pl = meta[idx]["predicted_label"][:11]
        cells = []
        for i in range(n_types):
            wt_v = gt_wt.get(idx, [0]*n_types)[i]
            ev_v = gt_ev.get(idx, [0]*n_types)[i]
            gp_v = preds.get(idx, [0]*n_types)[i]
            vals = [wt_v, ev_v, gp_v]
            mark = lambda v: "1" if v else "·"
            flag = "*" if len(set(vals)) > 1 else " "
            cells.append(f"{mark(wt_v)}  {mark(ev_v)}  {mark(gp_v)}{flag}")
        row = f"{idx:>4}  {tl:>11}  {pl:>11}  " + "  ".join(f"{c:^9}" for c in cells)
        print(row)

    print(sep)
    print("\nLegend:  W=WT  E=Evelyn  G=GPT-4o  1=annotated  ·=not annotated  *=disagreement")
    print("Types:   " + "  ".join(f"{s}={t[:20]}" for s, (t, _, _) in zip(SHORT_LABELS, TYPES)))
    print()


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--pred", default="test_gpt4o_validation_38.csv")
    parser.add_argument("--gt",   default="error_analysis_common_agree.xlsx")
    parser.add_argument("--out",  default="annotation_comparison",
                        help="Output file base name (no extension)")
    args = parser.parse_args()

    for p in (args.pred, args.gt):
        if not os.path.exists(p):
            print(f"File not found: {p}")
            sys.exit(1)

    preds, meta = load_predictions(args.pred)
    gt_wt, gt_ev = load_gt(args.gt, valid_indices=set(preds.keys()))

    idxs = sorted(set(preds.keys()) & set(gt_wt.keys()))
    print(f"\nSamples: {len(idxs)}")

    print_terminal(idxs, gt_wt, gt_ev, preds, meta)

    out_dir = os.path.dirname(args.pred) or "."
    build_excel(idxs, gt_wt, gt_ev, preds, meta,
                os.path.join(out_dir, args.out + ".xlsx"))
    build_csv(   idxs, gt_wt, gt_ev, preds, meta,
                os.path.join(out_dir, args.out + ".csv"))


if __name__ == "__main__":
    main()
