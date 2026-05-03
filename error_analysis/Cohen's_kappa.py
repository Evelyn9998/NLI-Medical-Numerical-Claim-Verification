"""
Inter-Annotator Agreement Analysis
===================================
Computes per-error-type counts, agreement statistics, and Cohen's kappa
for two annotators (wt and Evelyn) on the Llama-3.1-8B-Instruct error analysis dataset.

Dependencies:
    pip install pandas openpyxl scikit-learn numpy
"""

import pandas as pd
import numpy as np
from sklearn.metrics import cohen_kappa_score
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Configuration ─────────────────────────────────────────────────────────────

FILE_WT     = "error_analysis_50_wt_1.xlsx"
FILE_EVELYN = "error_analysis_50_Evelyn_1.xlsx"
OUTPUT_FILE = "Cohen's_kappa_result.xlsx"

ANN_COLS = [
    "failed parameter extraction",
    "verification logic errors",
    "incorrect formula & criteria application",
    "computation errors",
    "omitting the calculation process or result",
    "other errors",
]

DISPLAY_NAMES = [
    "failed_parameter_extraction",
    "verification_logic_error",
    "incorrect_formula&criteria_application",
    "computation_error",
    "omitted_calculation_process_or_result",
    "other_error",
]


# ── Data Loading ──────────────────────────────────────────────────────────────

def load_annotations(path: str) -> pd.DataFrame:
    """Load annotation file and drop the totals row (NaN index)."""
    df = pd.read_excel(path)
    df = df[df["index"].notna()].reset_index(drop=True)
    return df


# ── Statistics ────────────────────────────────────────────────────────────────

def compute_stats(wt: pd.DataFrame, ev: pd.DataFrame) -> list[dict]:
    """Compute per-error-type agreement statistics and Cohen's kappa."""
    n = len(wt)
    results = []
    for col, dname in zip(ANN_COLS, DISPLAY_NAMES):
        w = wt[col].astype(int)
        e = ev[col].astype(int)
        both_1 = int(((w == 1) & (e == 1)).sum())
        both_0 = int(((w == 0) & (e == 0)).sum())
        w1e0   = int(((w == 1) & (e == 0)).sum())
        w0e1   = int(((w == 0) & (e == 1)).sum())
        kappa  = cohen_kappa_score(w, e)
        results.append({
            "name":    dname,
            "wt":      int(w.sum()),
            "ev":      int(e.sum()),
            "both1":   both_1,
            "both0":   both_0,
            "w1e0":    w1e0,
            "w0e1":    w0e1,
            "disagree": n - both_1 - both_0,
            "kappa":   round(kappa, 3),
            "n":       n,
        })
    return results


def interpret_kappa(k: float) -> str:
    if k >= 0.80:  return "Almost Perfect"
    if k >= 0.60:  return "Substantial"
    if k >= 0.40:  return "Moderate"
    if k >= 0.20:  return "Fair"
    return "Poor"


def print_summary(rows: list[dict]) -> None:
    print(f"\n{'='*80}")
    print(f"{'Error Type':<45} {'wt':>5} {'Ev':>5} {'Kappa':>7}  Interpretation")
    print(f"{'-'*80}")
    for r in rows:
        print(
            f"{r['name']:<45} {r['wt']:>5} {r['ev']:>5} "
            f"{r['kappa']:>7.3f}  {interpret_kappa(r['kappa'])}"
        )
    avg_k = round(np.mean([r["kappa"] for r in rows]), 3)
    print(f"{'-'*80}")
    print(f"{'Average Cohen\'s κ':<45} {'':>5} {'':>5} {avg_k:>7.3f}")
    print(f"{'='*80}\n")


# ── Excel Output ──────────────────────────────────────────────────────────────

NAVY   = "1E2761"
ICE    = "CADCFC"
WHITE  = "FFFFFF"
LGRAY  = "F4F6FB"
GREEN  = "D5E8D4"
YELLOW = "FFF2CC"
RED    = "F8CECC"
ORANGE = "FFE6CC"


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _hc(ws, r, c, v, bg=NAVY, fg=WHITE, bold=True, left=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, color=fg, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(
        horizontal="left" if left else "center",
        vertical="center", wrap_text=True
    )
    return cell


def _dc(ws, r, c, v, bg=WHITE, bold=False, color="1A1A1A", left=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=bold, name="Arial", size=10, color=color)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(
        horizontal="left" if left else "center",
        vertical="center"
    )
    return cell


def _kappa_colors(k: float):
    if k >= 0.80:   return "2563EB", "DBEAFE"
    elif k >= 0.60: return "16A34A", "D5E8D4"
    elif k >= 0.40: return "D97706", "FFF2CC"
    else:           return "DC2626", "F8CECC"


def save_excel(rows: list[dict], output_path: str) -> None:
    bdr = _thin_border()
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Annotation Agreement"

    NCOLS = 9  # A–I

    # Title rows
    ws.merge_cells(f"A1:I1")
    ws["A1"].value = (
        "Inter-Annotator Agreement  —  "
        "Llama-3.1-8B-Instruct Error Analysis (n=50)"
    )
    ws["A1"].font      = Font(bold=True, name="Arial", size=13, color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:I2")
    ws["A2"].value = (
        "Annotators: wt  &  Evelyn   |   "
        "Samples: 50   |   Method: Zero-Shot CoT"
    )
    ws["A2"].font      = Font(name="Arial", size=10, color="555555", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # Column headers (Agreement column removed)
    hdrs = [
        "Error Type", "wt\n(Count)", "Evelyn\n(Count)",
        "Both=1\n(Agree+)", "Both=0\n(Agree−)",
        "wt=1\nEvelyn=0", "wt=0\nEvelyn=1",
        "Disagreement\n(n/50)", "Cohen's κ",
    ]
    for ci, h in enumerate(hdrs, 1):
        _hc(ws, 3, ci, h)
    ws.row_dimensions[3].height = 36

    # Data rows
    for ri, row in enumerate(rows, 4):
        bg = WHITE if ri % 2 == 0 else LGRAY
        _dc(ws, ri, 1, row["name"],    bg=bg, left=True)
        _dc(ws, ri, 2, row["wt"],      bg=bg, bold=True, color=NAVY)
        _dc(ws, ri, 3, row["ev"],      bg=bg, bold=True, color=NAVY)
        _dc(ws, ri, 4, row["both1"],   bg=GREEN)
        _dc(ws, ri, 5, row["both0"],   bg=GREEN)
        _dc(ws, ri, 6, row["w1e0"],    bg=ORANGE if row["w1e0"] > 0 else bg)
        _dc(ws, ri, 7, row["w0e1"],    bg=ORANGE if row["w0e1"] > 0 else bg)

        dis_bg = (RED if row["disagree"] > 10
                  else (YELLOW if row["disagree"] > 5 else GREEN))
        _dc(ws, ri, 8, f"{row['disagree']} / {row['n']}", bg=dis_bg)

        kb, kbg = _kappa_colors(row["kappa"])
        _dc(ws, ri, 9, row["kappa"], bg=kbg, bold=True, color=kb)
        ws.row_dimensions[ri].height = 22

    # Totals row
    tr = len(rows) + 4
    _hc(ws, tr, 1, "Total / Avg κ", left=True)
    _dc(ws, tr, 2, sum(r["wt"]    for r in rows), bg=NAVY, bold=True, color=WHITE)
    _dc(ws, tr, 3, sum(r["ev"]    for r in rows), bg=NAVY, bold=True, color=WHITE)
    _dc(ws, tr, 4, sum(r["both1"] for r in rows), bg=NAVY, bold=True, color=WHITE)
    _dc(ws, tr, 5, sum(r["both0"] for r in rows), bg=NAVY, bold=True, color=WHITE)
    _dc(ws, tr, 6, sum(r["w1e0"]  for r in rows), bg=NAVY, bold=True, color=WHITE)
    _dc(ws, tr, 7, sum(r["w0e1"]  for r in rows), bg=NAVY, bold=True, color=WHITE)
    _dc(ws, tr, 8, "", bg=NAVY)
    avg_k = round(np.mean([r["kappa"] for r in rows]), 3)
    _dc(ws, tr, 9, avg_k, bg=NAVY, bold=True, color=ICE)
    ws.row_dimensions[tr].height = 22

    # Legend
    lr = tr + 2
    ws.merge_cells(f"A{lr}:I{lr}")
    c = ws.cell(row=lr, column=1)
    c.value = (
        "Cohen's κ:  ≥0.80 Almost Perfect (blue)  |  "
        "0.60–0.79 Substantial (green)  |  "
        "0.40–0.59 Moderate (yellow)  |  "
        "<0.40 Fair/Poor (red)"
    )
    c.font      = Font(name="Arial", size=9, italic=True, color="555555")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[lr].height = 16

    # Borders
    for row in ws.iter_rows(min_row=3, max_row=tr, min_col=1, max_col=NCOLS):
        for cell in row:
            cell.border = bdr

    # Column widths
    for i, w in enumerate([44, 10, 12, 11, 11, 12, 12, 14, 11], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    wb.save(output_path)
    print(f"Saved: {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    wt_df = load_annotations(FILE_WT)
    ev_df = load_annotations(FILE_EVELYN)

    assert len(wt_df) == len(ev_df), "Files have different number of samples"
    assert (wt_df["index"].values == ev_df["index"].values).all(), \
        "Sample indices do not match between the two files"

    stats = compute_stats(wt_df, ev_df)
    print_summary(stats)
    save_excel(stats, OUTPUT_FILE)
