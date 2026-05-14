"""
Evaluate LLM-as-judge results against human annotations.
=========================================================
Computes accuracy, precision, recall, F1, and Cohen's kappa
(overall + per error type) for three comparisons:
  - common_agree (both annotators) vs model
  - WT vs model
  - Evelyn vs model

Usage:
    python evaluate_results.py --pred test_gpt4o_validation_38.csv
    python evaluate_results.py --pred test_gpt4o_validation_38.csv --gt error_analysis_common_agree.xlsx
"""

import argparse
import csv
import os
import re
import sys

import openpyxl

TYPES = [
    "failed_parameter_extraction",
    "verification_logic_error",
    "incorrect_formula_criteria",
    "computation_error",
    "omitted_calculation_process_or_result",
    "other_error",
]
SHORT = ["t1", "t2", "t3", "t4", "t5", "t6"]

GT_COLS = [
    ("failed_parameter_extraction",           "Evelyn - failed parameter extraction",               "WT - failed parameter extraction"),
    ("verification_logic_error",              "Evelyn - verification logic errors ",                "WT - verification logic errors"),
    ("incorrect_formula_criteria",            "Evelyn - incorrect formula & criteria application",  "WT -incorrect formula & criteria application"),
    ("computation_error",                     "Evelyn - computation errors",                        "WT - computation errors"),
    ("omitted_calculation_process_or_result", "Evelyn - omitting the calculation process or result","WT - omitting the calculation process or result"),
    ("other_error",                           "Evelyn - other errors",                              "WT - other errors"),
]


def load_ground_truth(xlsx_path: str, valid_indices: set):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    gt_agree, gt_wt, gt_ev = {}, {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        try:
            idx = int(float(str(d["index"])))
        except Exception:
            continue
        if valid_indices and idx not in valid_indices:
            continue
        gt_agree[idx] = [
            1 if (int(d.get(ec) or 0) == 1 and int(d.get(wc) or 0) == 1) else 0
            for _, ec, wc in GT_COLS
        ]
        gt_wt[idx] = [int(d.get(wc) or 0) for _, ec, wc in GT_COLS]
        gt_ev[idx] = [int(d.get(ec) or 0) for _, ec, wc in GT_COLS]

    return gt_agree, gt_wt, gt_ev


def load_predictions(csv_path: str):
    preds = {}
    with open(csv_path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            try:
                idx = int(float(row["index"]))
            except Exception:
                continue
            preds[idx] = [int(row.get(t, 0)) for t in TYPES]
    return preds


def compute_kappa(g_list, p_list):
    n = len(g_list)
    if n == 0:
        return 0.0
    tp = sum(g == 1 and p == 1 for g, p in zip(g_list, p_list))
    tn = sum(g == 0 and p == 0 for g, p in zip(g_list, p_list))
    fp = sum(g == 0 and p == 1 for g, p in zip(g_list, p_list))
    fn = sum(g == 1 and p == 0 for g, p in zip(g_list, p_list))
    po = (tp + tn) / n
    pe = ((tp + fn) / n * (tp + fp) / n) + ((tn + fp) / n * (tn + fn) / n)
    return (po - pe) / (1 - pe) if (1 - pe) != 0 else 1.0


def evaluate(gt_dict: dict, pred_dict: dict):
    idxs = sorted(gt_dict.keys() & pred_dict.keys())
    if not idxs:
        return None

    TP = FP = FN = TN = 0
    all_g, all_p = [], []
    for idx in idxs:
        for g, p in zip(gt_dict[idx], pred_dict[idx]):
            if   g == 1 and p == 1: TP += 1
            elif g == 0 and p == 1: FP += 1
            elif g == 1 and p == 0: FN += 1
            else:                   TN += 1
            all_g.append(g)
            all_p.append(p)

    n     = len(all_g)
    prec  = TP / (TP + FP) if TP + FP else 0.0
    rec   = TP / (TP + FN) if TP + FN else 0.0

    per_type_kappa = []
    for i in range(len(TYPES)):
        g_l = [gt_dict[idx][i] for idx in idxs]
        p_l = [pred_dict[idx][i] for idx in idxs]
        per_type_kappa.append(compute_kappa(g_l, p_l))

    return {
        "n_samples": len(idxs),
        "n_decisions": n,
        "acc":   (TP + TN) / n,
        "prec":  prec,
        "rec":   rec,
        "f1":    2 * prec * rec / (prec + rec) if prec + rec else 0.0,
        "kappa": compute_kappa(all_g, all_p),
        "TP": TP, "FP": FP, "FN": FN, "TN": TN,
        "per_type_kappa": per_type_kappa,
    }


def print_tables(results: dict):
    labels  = list(results.keys())
    metrics = list(results.values())

    # ── Summary table ────────────────────────────────────────────────────────
    col_w   = [max(len(lb), 22) for lb in labels]
    row_hdr = 12

    header = f"  {'Metric':<{row_hdr}} │ " + " │ ".join(f"{lb:^{w}}" for lb, w in zip(labels, col_w))
    sep    = f"  {'─'*row_hdr}─┼─" + "─┼─".join("─" * w for w in col_w)

    print("\n── Overall Metrics " + "─" * (len(header) - 20))
    print(header)
    print(sep)

    def row(name, vals):
        return f"  {name:<{row_hdr}} │ " + " │ ".join(f"{v:^{w}}" for v, w in zip(vals, col_w))

    r = metrics
    print(row("Accuracy",  [f"{m['acc']:.1%}  ({m['TP']+m['TN']}/{m['n_decisions']})" for m in r]))
    print(row("Precision", [f"{m['prec']:.1%}" for m in r]))
    print(row("Recall",    [f"{m['rec']:.1%}"  for m in r]))
    print(row("F1",        [f"{m['f1']:.3f}"   for m in r]))
    print(row("Kappa",     [f"{m['kappa']:.3f}" for m in r]))
    print(sep)
    print(row("TP / FP",   [f"{m['TP']} / {m['FP']}" for m in r]))
    print(row("FN / TN",   [f"{m['FN']} / {m['TN']}" for m in r]))
    print(row("Samples",   [f"n={m['n_samples']}" for m in r]))

    # ── Per-type kappa table ─────────────────────────────────────────────────
    type_hdr = 46
    kappa_w  = [max(len(lb), 7) for lb in labels]

    print(f"\n── Per-type Cohen's Kappa " + "─" * max(0, len(header) - 26))
    th = f"  {'':4} {'Error Type':<{type_hdr}} " + " ".join(f"{lb:>{w}}" for lb, w in zip(labels, kappa_w))
    print(th)
    print("  " + "─" * (len(th) - 2))
    for i, (s, t) in enumerate(zip(SHORT, TYPES)):
        kappas = " ".join(f"{m['per_type_kappa'][i]:>{w}.3f}" for m, w in zip(metrics, kappa_w))
        print(f"  {s:<4} {t:<{type_hdr}} {kappas}")
    print()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--pred", required=True,
                        help="Prediction CSV (output of ea_llm_judge.py)")
    parser.add_argument("--gt",   default="error_analysis_common_agree.xlsx",
                        help="Ground truth xlsx (default: error_analysis_common_agree.xlsx)")
    parser.add_argument("--model-name", default="",
                        help="Model label shown in output (auto-detected from filename if omitted)")
    args = parser.parse_args()

    stem = os.path.splitext(os.path.basename(args.pred))[0]
    m = re.match(r"error_fewshot_(.+?)_\d{8}_\d{6}$", stem)
    model_label = args.model_name or (m.group(1) if m else stem)

    for path in (args.pred, args.gt):
        if not os.path.exists(path):
            print(f"File not found: {path}")
            sys.exit(1)

    preds = load_predictions(args.pred)
    pred_indices = set(preds.keys())
    gt_agree, gt_wt, gt_ev = load_ground_truth(args.gt, valid_indices=pred_indices)

    print(f"\nPrediction file : {args.pred}")
    print(f"Ground truth    : {args.gt}")
    print(f"Samples in pred : {len(preds)}  |  matched in GT: {len(gt_agree)}")

    results = {
        f"common_agree vs {model_label}": evaluate(gt_agree, preds),
        f"WT vs {model_label}":           evaluate(gt_wt,    preds),
        f"Evelyn vs {model_label}":       evaluate(gt_ev,    preds),
    }
    results = {k: v for k, v in results.items() if v is not None}
    print_tables(results)


if __name__ == "__main__":
    main()
